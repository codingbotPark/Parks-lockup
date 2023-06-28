import face_recognition
from openpyxl import Workbook
import os
import numpy as np
import util

# 로드할 파일의 파일확장자
fileExtension = [
    '.jpg',
    '.png',
    '.jpeg',
]
imgDirPath = "./imgs"
# registImgs 는 __init__.py 에 의해 호출된다, 이에따른 경로를 설정
def loadFile(file,dir):
    # .jpg, .png, .jpeg 파일만 로드
    if '.' + file.split('.')[-1] in fileExtension:
        return os.path.join(dir,file)
    

# --------

def registImgs():
    registerList = list(filter(None,map(lambda file:loadFile(file,imgDirPath),os.listdir(imgDirPath))))
    encodedUsers= list(filter(None,list(map(lambda path:encodingImg(path,registerList),registerList))))
    makeExel(encodedUsers)
    

def encodingImg(path,registerList):
    loadedImg = face_recognition.load_image_file(path)
    EncodedImg = face_recognition.face_encodings(loadedImg)

    if (len(EncodedImg)):
        return {
            'encodedImg':EncodedImg[0],
            'name':util.getFileNameFromPath(path)
        }
    else:
        print("Cannot find face")
        return None



def makeExel(encodedUsers):
    workbook = Workbook()
    # 기본 시트 선택

    ws = workbook.active

    # 이름부터 넣는다
    for row_idx, row in enumerate(encodedUsers,start=1):
        ws.cell(row=row_idx, column=1,value=row['name'])

    # for row_idx, row in enumerate(encodedUsers,start=2):

    # # 인덱스와 값으로 사용한다
    for row_idx, row in enumerate(encodedUsers, start=1):
        for col_idx, value in enumerate(row['encodedImg'], start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)
        print(row['name'],"등록")
    workbook.save('user.xlsx')
    

